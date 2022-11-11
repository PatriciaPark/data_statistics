import re
from django.shortcuts import render, redirect
from datetime import datetime as dt
from .models import *
# from signup.models import *
from django.http import HttpResponse
from django.core.files.storage import FileSystemStorage
import openpyxl 
from openpyxl.utils.dataframe import dataframe_to_rows
from pandas import DataFrame
from .models import tbl_invoice_daily, tbl_product, tbl_store
import random
# from django.shortcuts import render,redirect
# from http.client import HTTPResponse
# from .models import SampleModel
# import json
# import openpyxl

# Create your views here.
def dailyToDB(request):
    # try:
    if request.method == 'POST' and request.FILES['fileInput']:
        file = request.FILES['fileInput']
        fs = FileSystemStorage()
        filename = fs.save(file.name, file)
        uploaded_file_url = fs.url(filename)
        excel_file = uploaded_file_url
        wb = openpyxl.load_workbook("."+excel_file, data_only=True)
        ws = wb.worksheets[1]
        grade_dic = {}
        # Converting a worksheet to a Dataframe
        data = ws.values

        # 필드명(첫행) 뽑아내어 cols에 넣어주고 뽑아낸 행은 data에서 삭제된다.
        # 여기서 next()는 한행 한행 iter 해주는 역할을 하고
        # [0:]은 필드를 col기준 어디서부터 선택할 것인지이다. 예)[1:]는 두번째 col부터 선택

        cols = next(data)[0:]
        cols = next(data)[0:]
        cols = next(data)[0:]
        cols = next(data)[0:]
        cols = next(data)[0:]
        df = DataFrame(data, columns=cols)
        
        date_day    = str(filename[3:5].strip())
        print("int(filename[3:5]): ", date_day)
        date_month  = str(filename[:2].strip())
        print("int(filename[0:2]): ", date_month)
        date_year = str(int(str(ws.cell(row=2,column=1).value).split("：")[1].split("年")[0].strip())+1911)
        print(date_year)

        date_y_m_d = date_year + "-" + date_month + "-" + date_day
        print("date_y_m_d: ", date_y_m_d)    
        fromdate_time_obj = dt.strptime(date_y_m_d, '%Y-%m-%d')
        print("fromdate_time_obj: ",fromdate_time_obj)
        #format_data = "%d/%m/%y %H:%M:%S.%f"

        for dbframe in df.itertuples():

            product = tbl_product.objects.get(product_code_bystore = str(dbframe.貨號))
            store = tbl_store.objects.get(store_code = str(dbframe.門市代號))
      
            obj = tbl_invoice_daily.objects.create(
                invoice_day_date    = fromdate_time_obj,
                invoice_day_save    = dbframe.上存量,
                invoice_day_buy     = dbframe.進貨量,
                invoice_day_return  = dbframe.退貨量,
                invoice_day_sale    = dbframe.銷貨量,
                invoice_day_stock   = dbframe.庫存量,
                invoice_day_product = product,
                invoice_day_store = store
                )
            obj.save()

    # except:
    #     message = 'An unknown error occurred.'
    #     return render(request, 'main/error.html',{"message": message})
        
    return redirect('/index/')

def monthlyToDB(request):
    return HttpResponse("upload to db monthlyExcel File")

def yearlyToDB(request):
    return HttpResponse("upload to db YearlyExcel File")

def storeInfoToDB(request):

    try:
        if request.method == 'POST' and request.FILES['fileInput']:
            file = request.FILES['fileInput']
            fs = FileSystemStorage()
            filename = fs.save(file.name, file)
            uploaded_file_url = fs.url(filename)
            excel_file = uploaded_file_url
            wb = openpyxl.load_workbook("."+excel_file, data_only=True)
            ws = wb.worksheets[1]
            grade_dic = {}
            # Converting a worksheet to a Dataframe
            data = ws.values

            # 필드명(첫행) 뽑아내어 cols에 넣어주고 뽑아낸 행은 data에서 삭제된다.
            # 여기서 next()는 한행 한행 iter 해주는 역할을 하고
            # [0:]은 필드를 col기준 어디서부터 선택할 것인지이다. 예)[1:]는 두번째 col부터 선택

            cols = next(data)[0:]
            cols = next(data)[0:]
            cols = next(data)[0:]
            cols = next(data)[0:]
            cols = next(data)[0:]
            df = DataFrame(data, columns=cols)
            
            date_day    = str(filename[3:5].strip())
            date_month  = str(filename[:2].strip())
            date_year = str(int(str(ws.cell(row=2,column=1).value).split("：")[1].split("年")[0].strip())+1911)
            date_y_m_d = date_year + "-" + date_month + "-" + date_day
        
            fromdate_time_obj = dt.strptime(date_y_m_d, '%Y-%m-%d')

            for dbframe in df.itertuples():

                obj = tbl_store.objects.create(
                    store_code      = dbframe.門市代號,
                    store_name      = dbframe.門市名稱,
                    store_district  = "東部",
                    store_city      = "七頭",
                    store_address   = "東部 七頭市",
                    store_detail   = "...",
                    store_start = fromdate_time_obj,
                    )
                obj.save()

    except:
        message = 'An unknown error occurred.'
        return render(request, 'main/error.html',{"message": message})
        
    return redirect('/index/')



# class ExcelUploadView(Views):
# def list(request):
#     excels = SampleModel.objects.all().order_by('id')
#     return render(request, 'index2.html', {'excels':excels})


# def post(self, request):
#         excelFile = request.FILES['file']

#         excel = openpyxl.load_workbook(excelFile, data_only=True)
#         work_sheet = excel.worksheets[0]

#         all_values = []
#         for row in work_sheet.rows:
#             row_value = []
#             for cell in row:
#                 row_value.append(cell.value)
#             all_values.append(row_value)

#         for row in all_values:
#             sample_model = SampleModel(Number=row[0], Name=row[1], Item=row[2])
#             sample_model.save()

#         response = {'status': 1, 'message': '엑셀파일이 정상적으로 업로드 됐습니다.'}
#         return HTTPResponse(json.dumps(response), content_type='application/json')