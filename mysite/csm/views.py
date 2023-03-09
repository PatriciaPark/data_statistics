#-*-coding: utf-8
from django.shortcuts import render, redirect
from django.http import JsonResponse
from .models import *
from django.core.files.storage import FileSystemStorage
from django.core.exceptions import ObjectDoesNotExist
from django.db import IntegrityError
from django.contrib import messages
from django.db.models import Sum
from datetime import datetime
from pandas import DataFrame
import openpyxl, calendar


# index
def index(request):
    # html에서 선택한 날짜 받아오기
    try:
        date  = request.GET.get('input-month')  #2022-11
        year  = date[:4]                        #2022
        month = date[5:]                        #11
    except TypeError:
        year  = datetime.today().year
        month = datetime.today().strftime('%m')
        
    # 불러온 달의 마지막일 (eg. 28일/30일/31일)
    last_date = calendar.monthrange(int(year), int(month))[1]

    # get select box value
    getloc   = request.GET.get('add-loc')
    getCity  = request.GET.get('add-city')
    getStore = request.GET.get('add-str')   
    
    # set select box value
    loc  = Store.objects.values('str_loc').distinct()
    city = Store.objects.filter(str_loc=getloc).values('str_city').distinct()
    str  = Store.objects.filter(str_loc=getloc, str_city=getCity).values('str_name').order_by('str_name')
    
    # select box list
    product = Product.objects.all()
    
    # html에서 선택한 select box value 받아오기 및 정렬
    prdcode = request.GET.get('input-prd')  #상품코드
    sort    = request.GET.get('input-sort') #정렬
    
    # read data
    # 지역만 선택
    if getloc != 'all' and (getCity is None or getCity == 'all') and (getStore is None or getStore == 'all'):
        # select box 선택에 따른 정렬
        if prdcode:
            if sort == 'strname': 
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('str_code__str_name', 'sum_m_code','prd_code')
            elif  sort == 'prdSave':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_save','sum_m_code','prd_code')
            elif  sort == 'prdBuy':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_buy','sum_m_code','prd_code')
            elif  sort == 'prdReturn':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_return','sum_m_code','prd_code')
            elif  sort == 'prdSale':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_sale','sum_m_code','prd_code')
            elif  sort == 'prdStock':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_stock','sum_m_code','prd_code')
            elif sort == 'prdStockMon':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_stock_mon','sum_m_code','prd_code')
            else:
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('sum_m_code','prd_code')
            #total
            sumSave   = SumMonthly.objects.filter(str_code__str_loc=getloc, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_save'))
            sumBuy    = SumMonthly.objects.filter(str_code__str_loc=getloc, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_buy'))
            sumReturn = SumMonthly.objects.filter(str_code__str_loc=getloc, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_return'))
            sumSale   = SumMonthly.objects.filter(str_code__str_loc=getloc, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_sale'))
            sumStock  = SumMonthly.objects.filter(str_code__str_loc=getloc, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_stock'))
            try:
                ttlSumDl = list(sumSale.values())[0]/last_date # 每日銷貨總計= 銷貨量/當月總天數
                ttlerr   = list(sumSale.values())[0]/ttlSumDl  # 誤差值= 銷貨量/每日銷貨總計
            except:
                ttlSumDl = '-'
                ttlerr   = '-'
        else:
            if sort == 'strname': 
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('str_code__str_name', 'sum_m_code','prd_code')
            elif  sort == 'prdSave':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_save','sum_m_code','prd_code')
            elif  sort == 'prdBuy':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_buy','sum_m_code','prd_code')
            elif  sort == 'prdReturn':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_return','sum_m_code','prd_code')
            elif  sort == 'prdSale':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_sale','sum_m_code','prd_code')
            elif  sort == 'prdStock':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_stock','sum_m_code','prd_code')
            elif sort == 'prdStockMon':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_stock_mon','sum_m_code','prd_code')
            else:
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('sum_m_code','prd_code')
            #total
            sumSave   = SumMonthly.objects.filter(str_code__str_loc=getloc, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_save'))
            sumBuy    = SumMonthly.objects.filter(str_code__str_loc=getloc, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_buy'))
            sumReturn = SumMonthly.objects.filter(str_code__str_loc=getloc, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_return'))
            sumSale   = SumMonthly.objects.filter(str_code__str_loc=getloc, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_sale'))
            sumStock  = SumMonthly.objects.filter(str_code__str_loc=getloc, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_stock'))
            try:
                ttlSumDl = list(sumSale.values())[0]/last_date # 每日銷貨總計= 銷貨量/當月總天數
                ttlerr   = list(sumSale.values())[0]/ttlSumDl  # 誤差值= 銷貨量/每日銷貨總計
            except:
                ttlSumDl = '-'
                ttlerr   = '-'
    # 지역 및 시까지 선택
    elif getloc != 'all' and getCity is not None and (getStore is None or getStore == 'all'):
        # select box 선택에 따른 정렬
        if prdcode:
            if sort == 'strname': 
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('str_code__str_name', 'sum_m_code','prd_code')
            elif  sort == 'prdSave':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_save','sum_m_code','prd_code')
            elif  sort == 'prdBuy':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_buy','sum_m_code','prd_code')
            elif  sort == 'prdReturn':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_return','sum_m_code','prd_code')
            elif  sort == 'prdSale':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_sale','sum_m_code','prd_code')
            elif  sort == 'prdStock':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_stock','sum_m_code','prd_code')
            elif sort == 'prdStockMon':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_stock_mon','sum_m_code','prd_code')
            else:
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('sum_m_code','prd_code')
            #total
            sumSave   = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_save'))
            sumBuy    = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_buy'))
            sumReturn = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_return'))
            sumSale   = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_sale'))
            sumStock  = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_stock'))
            try:
                ttlSumDl = list(sumSale.values())[0]/last_date # 每日銷貨總計= 銷貨量/當月總天數
                ttlerr   = list(sumSale.values())[0]/ttlSumDl  # 誤差值= 銷貨量/每日銷貨總計
            except:
                ttlSumDl = '-'
                ttlerr   = '-'
        else:
            if sort == 'strname': 
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('str_code__str_name', 'sum_m_code','prd_code')
            elif  sort == 'prdSave':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_save','sum_m_code','prd_code')
            elif  sort == 'prdBuy':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_buy','sum_m_code','prd_code')
            elif  sort == 'prdReturn':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_return','sum_m_code','prd_code')
            elif  sort == 'prdSale':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_sale','sum_m_code','prd_code')
            elif  sort == 'prdStock':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_stock','sum_m_code','prd_code')
            elif sort == 'prdStockMon':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_stock_mon','sum_m_code','prd_code')
            else:
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('sum_m_code','prd_code')
            #total
            sumSave   = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_save'))
            sumBuy    = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_buy'))
            sumReturn = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_return'))
            sumSale   = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_sale'))
            sumStock  = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_stock'))
            try:
                ttlSumDl = list(sumSale.values())[0]/last_date # 每日銷貨總計= 銷貨量/當月總天數
                ttlerr   = list(sumSale.values())[0]/ttlSumDl  # 誤差值= 銷貨量/每日銷貨總計
            except:
                ttlSumDl = '-'
                ttlerr   = '-'
    #지역, 시, 매장 선택
    elif getloc != 'all' and getCity is not None and getStore is not None:
        # select box 선택에 따른 정렬
        if prdcode:
            if sort == 'strname': 
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, str_code__str_name=getStore, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('str_code__str_name', 'sum_m_code','prd_code')
            elif  sort == 'prdSave':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, str_code__str_name=getStore, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_save','sum_m_code','prd_code')
            elif  sort == 'prdBuy':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, str_code__str_name=getStore, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_buy','sum_m_code','prd_code')
            elif  sort == 'prdReturn':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, str_code__str_name=getStore, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_return','sum_m_code','prd_code')
            elif  sort == 'prdSale':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, str_code__str_name=getStore, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_sale','sum_m_code','prd_code')
            elif  sort == 'prdStock':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, str_code__str_name=getStore, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_stock','sum_m_code','prd_code')
            elif sort == 'prdStockMon':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, str_code__str_name=getStore, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_stock_mon','sum_m_code','prd_code')
            else:
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, str_code__str_name=getStore, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('sum_m_code','prd_code')
            #total
            sumSave   = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, str_code__str_name=getStore, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_save'))
            sumBuy    = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, str_code__str_name=getStore, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_buy'))
            sumReturn = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, str_code__str_name=getStore, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_return'))
            sumSale   = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, str_code__str_name=getStore, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_sale'))
            sumStock  = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, str_code__str_name=getStore, prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_stock'))
            try:
                ttlSumDl = list(sumSale.values())[0]/last_date # 每日銷貨總計= 銷貨量/當月總天數
                ttlerr   = list(sumSale.values())[0]/ttlSumDl  # 誤差值= 銷貨量/每日銷貨總計
            except:
                ttlSumDl = '-'
                ttlerr   = '-'
        else:
            if sort == 'strname': 
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, str_code__str_name=getStore, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('str_code__str_name', 'sum_m_code','prd_code')
            elif  sort == 'prdSave':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, str_code__str_name=getStore, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_save','sum_m_code','prd_code')
            elif  sort == 'prdBuy':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, str_code__str_name=getStore, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_buy','sum_m_code','prd_code')
            elif  sort == 'prdReturn':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, str_code__str_name=getStore, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_return','sum_m_code','prd_code')
            elif  sort == 'prdSale':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, str_code__str_name=getStore, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_sale','sum_m_code','prd_code')
            elif  sort == 'prdStock':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, str_code__str_name=getStore, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_stock','sum_m_code','prd_code')
            elif sort == 'prdStockMon':
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, str_code__str_name=getStore, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_stock_mon','sum_m_code','prd_code')
            else:
                monthData = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, str_code__str_name=getStore, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('sum_m_code','prd_code')
            #total
            sumSave   = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, str_code__str_name=getStore, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_save'))
            sumBuy    = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, str_code__str_name=getStore, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_buy'))
            sumReturn = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, str_code__str_name=getStore, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_return'))
            sumSale   = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, str_code__str_name=getStore, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_sale'))
            sumStock  = SumMonthly.objects.filter(str_code__str_loc=getloc, str_code__str_city=getCity, str_code__str_name=getStore, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_stock'))
            try:
                ttlSumDl = list(sumSale.values())[0]/last_date # 每日銷貨總計= 銷貨量/當月總天數
                ttlerr   = list(sumSale.values())[0]/ttlSumDl  # 誤差值= 銷貨量/每日銷貨總計
            except:
                ttlSumDl = '-'
                ttlerr   = '-'
    else:
        # select box 선택에 따른 정렬
        if prdcode:
            if sort == 'strname': 
                monthData = SumMonthly.objects.filter(prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('str_code__str_name', 'sum_m_code','prd_code')
            elif  sort == 'prdSave':
                monthData = SumMonthly.objects.filter(prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_save','sum_m_code','prd_code')
            elif  sort == 'prdBuy':
                monthData = SumMonthly.objects.filter(prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_buy','sum_m_code','prd_code')
            elif  sort == 'prdReturn':
                monthData = SumMonthly.objects.filter(prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_return','sum_m_code','prd_code')
            elif  sort == 'prdSale':
                monthData = SumMonthly.objects.filter(prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_sale','sum_m_code','prd_code')
            elif  sort == 'prdStock':
                monthData = SumMonthly.objects.filter(prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_stock','sum_m_code','prd_code')
            elif sort == 'prdStockMon':
                monthData = SumMonthly.objects.filter(prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_stock_mon','sum_m_code','prd_code')
            else:
                monthData = SumMonthly.objects.filter(prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('sum_m_code','prd_code')
            #total
            sumSave   = SumMonthly.objects.filter(prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_save'))
            sumBuy    = SumMonthly.objects.filter(prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_buy'))
            sumReturn = SumMonthly.objects.filter(prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_return'))
            sumSale   = SumMonthly.objects.filter(prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_sale'))
            sumStock  = SumMonthly.objects.filter(prd_code = prdcode, sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_stock'))
            try:
                ttlSumDl = list(sumSale.values())[0]/last_date # 每日銷貨總計= 銷貨量/當月總天數
                ttlerr   = list(sumSale.values())[0]/ttlSumDl  # 誤差值= 銷貨量/每日銷貨總計
            except:
                ttlSumDl = '-'
                ttlerr   = '-'
        else:
            if sort == 'strname': 
                monthData = SumMonthly.objects.filter(sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('str_code__str_name', 'sum_m_code','prd_code')
            elif  sort == 'prdSave':
                monthData = SumMonthly.objects.filter(sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_save','sum_m_code','prd_code')
            elif  sort == 'prdBuy':
                monthData = SumMonthly.objects.filter(sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_buy','sum_m_code','prd_code')
            elif  sort == 'prdReturn':
                monthData = SumMonthly.objects.filter(sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_return','sum_m_code','prd_code')
            elif  sort == 'prdSale':
                monthData = SumMonthly.objects.filter(sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_sale','sum_m_code','prd_code')
            elif  sort == 'prdStock':
                monthData = SumMonthly.objects.filter(sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_stock','sum_m_code','prd_code')
            elif sort == 'prdStockMon':
                monthData = SumMonthly.objects.filter(sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('-sum_m_stock_mon','sum_m_code','prd_code')
            else:
                monthData = SumMonthly.objects.filter(sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').order_by('sum_m_code','prd_code')
            #total
            sumSave   = SumMonthly.objects.filter(sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_save'))
            sumBuy    = SumMonthly.objects.filter(sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_buy'))
            sumReturn = SumMonthly.objects.filter(sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_return'))
            sumSale   = SumMonthly.objects.filter(sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_sale'))
            sumStock  = SumMonthly.objects.filter(sum_m_date__year=int(year), sum_m_date__month=int(month)).select_related('prd_code','str_code').aggregate(Sum('sum_m_stock'))
            try:
                ttlSumDl = list(sumSale.values())[0]/last_date # 每日銷貨總計= 銷貨量/當月總天數
                ttlerr   = list(sumSale.values())[0]/ttlSumDl  # 誤差值= 銷貨量/每日銷貨總計
            except:
                ttlSumDl = '-'
                ttlerr   = '-'
            
    # html 화면에 출력할 데이터 담을 리스트                    
    context = {'product':product, 'year':year, 'month':month, 'monthData':monthData, 'prdcode':prdcode, 'sort':sort, 'loc':loc, 'city':city, 'str':str,
               'sumSave':sumSave, 'sumBuy':sumBuy, 'sumReturn':sumReturn, 'sumSale':sumSale, 'sumStock':sumStock, 'ttlSumDl':ttlSumDl, 'ttlerr':ttlerr}

    return render(request, 'csm/index.html', context)

# excel file upload
def upload(request):
    if request.method == 'POST' and request.FILES['fileInput']:
        # file directory
        file = request.FILES['fileInput']
        fs = FileSystemStorage()
        filename = fs.save(file.name.replace(' ' , ''), file)
        excel_file = '/media/'+ filename
        
        # get sheet name
        wb = openpyxl.load_workbook('.'+excel_file, data_only=True)
        sheetname = wb.sheetnames
        
        try:
            for i in range(0, len(sheetname)):    
                ws = wb.worksheets[i]
                
                # Converting a worksheet to a Dataframe
                data = ws.values
                cols = next(data)[0:]
                cols = next(data)[0:]
                cols = next(data)[0:]
                cols = next(data)[0:]
                cols = next(data)[0:]
                df = DataFrame(data, columns=cols)
                
                # 날짜 추출      
                date_month  = str(filename.split(".")[0].split('_')[0][-2:].strip())
                date_year = str(int(str(ws.cell(row=2,column=1).value).split("：")[1].split("年")[0].strip())+1911)
                date_y_m = date_year + "-" + date_month
                fromdate_time_obj = datetime.strptime(date_y_m, '%Y-%m')
                
                # 불러온 달의 마지막일 (eg. 28일/30일/31일)
                last_date = calendar.monthrange(int(date_year), int(date_month))[1]
                print("***Updated... ",fromdate_time_obj)
                
                # 데이터 DB 저장
                for dbframe in df.itertuples():
                    # 제품/매장 테이블에서 외래키 참조
                    product, store = 0, 0
                    
                    # Get or Save prd data to DB (tbl_product)
                    try:
                        product = Product.objects.get(prd_code = str(dbframe.貨號))
                    except ObjectDoesNotExist:
                        try:
                            product = Product.objects.create(
                                prd_code = str(dbframe.貨號),
                                prd_barcode = str(dbframe.條碼),
                                prd_name = dbframe._5
                            )
                            product.save()
                        except IntegrityError:
                            pass
                    
                    # Get or Save str data to DB (tbl_store)    
                    try:
                        store = Store.objects.get(str_code = str(dbframe.門市代號))
                    except ObjectDoesNotExist:
                        try:
                            store = Store.objects.create(
                                str_code = dbframe.門市代號,
                                str_city = dbframe.縣市,
                                str_loc  = dbframe.區域,
                                str_name = dbframe.門市名稱
                            )
                            store.save()
                        except IntegrityError:
                            pass
                    
                    # 일일 판매량, 오차율, 월재고율
                    m_sale  = wb.worksheets[i]['K4'].value # 판매총합 셀
                    dailySale, errorRate, stockMon = 0, 0, 0
                    # dailySale = float(format(m_sale/last_date,".2f"))
                    dailySale = float(format(dbframe.銷貨量/last_date,".2f"))
                    
                    try : 
                        errorRate = format(dbframe.銷貨量/dailySale,".2f")
                    except ZeroDivisionError:
                        errorRate = None
                    try : 
                        stockMon = format(dbframe.庫存量/dbframe.銷貨量,".2f")
                    except ZeroDivisionError:
                        stockMon = None
                        
                    # Save sum data to DB (tbl_sum_m)
                    obj, created = SumMonthly.objects.update_or_create(
                        sum_m_date  = fromdate_time_obj,
                        str_code = store,
                        prd_code = product,
                        defaults={
                            'sum_m_date'  : fromdate_time_obj,
                            'sum_m_save'  : dbframe.上存量,
                            'sum_m_buy'   : dbframe.進貨量,
                            'sum_m_return': dbframe.退貨量,
                            'sum_m_sale'  : dbframe.銷貨量,
                            'sum_m_stock' : dbframe.庫存量,
                            'prd_code'    : product,
                            'str_code'    : store,
                            'sum_m_sale_ttl' : dailySale,
                            'sum_m_sale_err' : errorRate,
                            'sum_m_stock_mon': stockMon
                        }
                    )
                    obj.save()
                
        except:
            messages.warning(request, 'Invalid file')
            return redirect('/csm/')
    
    return redirect('/csm/?input-month='+date_year+'-'+date_month)

# Add Field Select Box
def select(request):
    # get select box value
    getloc = request.GET.get('location')
    
    # set select box value
    city = Store.objects.filter(str_loc=getloc).values('str_city').distinct().order_by('str_city')
    
    return JsonResponse({'data': list(city)})

def select2(request):
    # get select box value
    getCity = request.GET.get('city')
    
    # set select box value
    store = Store.objects.filter(str_city=getCity).values('str_name','str_code').order_by('str_name')
    
    return JsonResponse({'data': list(store)})