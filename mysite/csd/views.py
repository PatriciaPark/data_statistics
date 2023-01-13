#-*-coding: utf-8
from django.shortcuts import render, redirect
from .models import *
from django.core.files.storage import FileSystemStorage
from django.core.exceptions import ObjectDoesNotExist
from django.db import IntegrityError
from django.contrib import messages
from datetime import datetime
from pandas import DataFrame
import openpyxl, calendar


# index
def index(request):
    # html에서 선택한 날짜 받아오기
    try:
        date = request.GET.get('input-month')   #2022-11
        year = date[:4]     #2022
        month = date[5:]    #11
    except TypeError:
        year = datetime.today().year
        month = datetime.today().strftime('%m')

    # DB에서 상품별 합계 데이터 받아오기
    prd11530035 = list(SumDaily.objects.filter(prd_code = 11530035, sum_d_date__year=int(year), sum_d_date__month=int(month)).select_related('prd_code').order_by('sum_d_date')) #正官庄活蔘２８Ｄ高麗蔘活 / 力飲１００ｍｌ＊１０瓶
    prd11060162 = list(SumDaily.objects.filter(prd_code = 11060162, sum_d_date__year=int(year), sum_d_date__month=int(month)).select_related('prd_code').order_by('sum_d_date')) #正官庄高麗蔘精ＥＶＥＲＹ / ＴＩＭＥ－秘１０ｍｌ＊２０入
    prd17010087 = list(SumDaily.objects.filter(prd_code = 17010087, sum_d_date__year=int(year), sum_d_date__month=int(month)).select_related('prd_code').order_by('sum_d_date')) #預購正官庄活蔘２８Ｄ高麗 / 蔘活力飲禮盒１００ｍｌ＊８入
    prd17010088 = list(SumDaily.objects.filter(prd_code = 17010088, sum_d_date__year=int(year), sum_d_date__month=int(month)).select_related('prd_code').order_by('sum_d_date')) #預購正官庄高麗蔘石榴飲 / ５０ｍｌ＊９入
    prd17010004 = list(SumDaily.objects.filter(prd_code = 17010004, sum_d_date__year=int(year), sum_d_date__month=int(month)).select_related('prd_code').order_by('sum_d_date')) #預購正官庄高麗蔘野櫻莓飲
    prd17010002 = list(SumDaily.objects.filter(prd_code = 17010002, sum_d_date__year=int(year), sum_d_date__month=int(month)).select_related('prd_code').order_by('sum_d_date')) #預購正官庄高麗蔘精ＥＶＥ / ＲＹＴＩＭＥ１０ｍｌ＊３０入
    prdlist     = list(SumDaily.objects.filter(prd_code = 12345678, sum_d_date__year=int(year), sum_d_date__month=int(month)).select_related('prd_code').order_by('sum_d_date'))
    
    # print('###prdlist ',prdlist)
    
    # html 화면에 출력할 데이터 담을 리스트                    
    context = {'month':month, 'prdlist':prdlist, 'prd11530035':prd11530035, 'prd11060162':prd11060162, 'prd17010087':prd17010087, 'prd17010088':prd17010088, 'prd17010004':prd17010004, 'prd17010002':prd17010002}
    
    return render(request, 'csd/index.html', context)

# excel file upload
def upload(request):
    if  request.method == 'POST' and request.FILES['fileInput']:
        file = request.FILES['fileInput']
        fs = FileSystemStorage()
        filename = fs.save(file.name.replace(' ' , ''), file)
        # uploaded_file_url = fs.url(filename)
        # excel_file = uploaded_file_url
        # 파일명에 한자가 들어가 있으면 위 코드 사용시 인코딩 문제로 에러 발생
        excel_file = '/media/'+ filename
        
        # 시트 이름 가져오기
        wb = openpyxl.load_workbook('.'+excel_file, data_only=True)
        ws = wb.worksheets[0]
        sheetname = wb.sheetnames
        
        # Converting a worksheet to a Dataframe
        data = ws.values
        # 필드명(첫행) 뽑아내어 cols에 넣어주고 뽑아낸 행은 data에서 삭제된다.
        # 여기서 next()는 한행 한행 iter 해주는 역할을 하고
        # [0:]은 필드를 col기준 어디서부터 선택할 것인지이다. 예)[1:]는 두번째 col부터 선택
        cols = next(data)[0:]
        cols = next(data)[0:]
        cols = next(data)[0:]
        cols = next(data)[0:]
        cols = next(data)[0:]
        df = DataFrame(data, columns=cols)
        
        try:
            for i in range(0, len(sheetname)):
                # 엑셀 sheet 이름에서 날짜 추출 
                date_month  = str(sheetname[i].split('.')[0])
                date_day    = str(sheetname[i].split('.')[1])
                # 추출한 날짜 형변환
                date_year = str(int(str(ws.cell(row=2,column=1).value).split('：')[1].split('年')[0].strip())+1911)
                date_y_m_d = date_year + '-' + date_month + '-' + date_day
                fromdate_time_obj = datetime.strptime(date_y_m_d, '%Y-%m-%d')
                
                last_date = calendar.monthrange(int(date_year), int(date_month))[1]
                print("***Updated... ",fromdate_time_obj)
                
                # 엑셀 합계 데이터 들어있는 셀 읽어오기 (F4~J4)
                d_save  = wb.worksheets[i]['F4'].value
                d_buy   = wb.worksheets[i]['G4'].value
                d_return= wb.worksheets[i]['H4'].value
                d_sale  = wb.worksheets[i]['I4'].value
                d_stock = wb.worksheets[i]['J4'].value            
                
                # 데이터 DB 저장            
                product, store = 0, 0
                for dbframe in df.itertuples():
                    # Get or Save prd data to DB (tbl_product)
                    try:
                        product = Product.objects.get(prd_code = str(dbframe.貨號))
                    except ObjectDoesNotExist:
                        try:
                            product = Product.objects.create(
                                prd_code = str(dbframe.貨號),
                                prd_barcode = str(dbframe.條碼),
                                prd_name = dbframe._5
                            )
                            product.save()
                        except IntegrityError:
                            pass
                    
                    # Get or Save str data to DB (tbl_store)    
                    try:
                        store = Store.objects.get(str_code = str(dbframe.門市代號))
                    except ObjectDoesNotExist:
                        try:
                            store = Store.objects.create(
                                str_code = dbframe.門市代號,
                                str_name = dbframe.門市名稱
                            )
                            store.save()
                        except IntegrityError:
                            pass
                    
                    # 일일 데이터 DB 저장 (tbl_invoice_d)
                    obj, created = InvoiceDaily.objects.update_or_create(
                        inv_d_date  = fromdate_time_obj,
                        prd_code    = product,
                        str_code    = store,
                        defaults={
                            'inv_d_date'  : fromdate_time_obj,
                            'inv_d_save'  : dbframe.上存量,
                            'inv_d_buy'   : dbframe.進貨量,
                            'inv_d_return': dbframe.退貨量,
                            'inv_d_sale'  : dbframe.銷貨量,
                            'inv_d_stock' : dbframe.庫存量,
                            'prd_code'    : product,
                            'str_code'    : store
                        }
                        )
                    obj.save()
                    
                # Save sum data to DB (tbl_sum_d)
                obj_sum, created = SumDaily.objects.update_or_create(
                    sum_d_date  = fromdate_time_obj,
                    prd_code = product,
                    defaults={
                        'sum_d_date'  : fromdate_time_obj,
                        'sum_d_save'  : d_save,
                        'sum_d_buy'   : d_buy,
                        'sum_d_return': d_return,
                        'sum_d_sale'  : d_sale,
                        'sum_d_stock' : d_stock,
                        'prd_code'    : product
                    }
                )
                obj_sum.save()
            
                # 엑셀 시트에 없는 날짜 추출 및 형변환
                all_date = []
                saved_date = []
                
                for j in range(last_date):
                    all_date.append(j+1)
                    
                for k in range(0, len(sheetname)):     
                    saved_date.append(int(sheetname[k].split('.')[1]))

                # missing date data
                none_date = list(set(all_date)-set(saved_date))
                
                for l in range(0,len(none_date)):
                    date_y_m_d_null = date_year + '-' + date_month + '-' + str(none_date[l])
                    fromdate_time_obj_null = datetime.strptime(date_y_m_d_null, '%Y-%m-%d')
                    
                    # Save missing date data to DB
                    obj_sum, created = SumDaily.objects.get_or_create(
                        sum_d_date  = fromdate_time_obj_null,
                        prd_code = product,
                        defaults={
                            'sum_d_date'  : fromdate_time_obj_null,
                            'sum_d_save'  : 0,
                            'sum_d_buy'   : 0,
                            'sum_d_return': 0,
                            'sum_d_sale'  : 0,
                            'sum_d_stock' : 0,
                            'prd_code'    : product
                        }
                    )
                    obj_sum.save()
        except:
            messages.warning(request, 'Invalid file')
            return redirect('/csd/')
                    
    return redirect('/csd/?input-month='+date_year+'-'+date_month)